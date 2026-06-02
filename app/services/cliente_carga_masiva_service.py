"""Carga masiva de clientes desde Excel (solo administrador)."""
from __future__ import annotations

import unicodedata
from io import BytesIO
from typing import Any

from email_validator import EmailNotValidError, validate_email
from fastapi import HTTPException, UploadFile, status
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.enums import RolUsuario
from app.models.cliente import Cliente
from app.models.usuario import Usuario
from app.schemas.cliente_carga import ClienteCargaMasivaResultado, FilaCargaClienteResultado

PLANTILLA_HEADERS = [
    "nombre",
    "apellidos",
    "nit",
    "direccion",
    "telefono",
    "correo",
    "ciudad",
    "vendedor",
]

_HEADER_ALIASES: dict[str, str] = {
    "nombre": "nombre",
    "apellidos": "apellidos",
    "nit": "nit",
    "direccion": "direccion",
    "dirección": "direccion",
    "telefono": "telefono",
    "teléfono": "telefono",
    "correo": "correo",
    "email": "correo",
    "ciudad": "ciudad",
    "vendedor": "vendedor",
    "vendedor_asignado": "vendedor",
    "nombre_vendedor": "vendedor",
}


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    return _HEADER_ALIASES.get(text, text)


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_correo(raw: str) -> str | None:
    if not raw:
        return None
    try:
        return validate_email(raw, check_deliverability=False).normalized
    except EmailNotValidError as exc:
        raise ValueError(f"Correo inválido: {raw}") from exc


def _resolve_vendedor_exacto(db: Session, nombre_vendedor: str) -> Usuario:
    nombre = nombre_vendedor.strip()
    if not nombre:
        raise ValueError("La columna 'vendedor' es obligatoria.")
    rows = list(
        db.execute(
            select(Usuario).where(
                Usuario.rol == RolUsuario.VENDEDOR.value,
                Usuario.activo.is_(True),
                Usuario.nombre == nombre,
            )
        ).scalars()
    )
    if not rows:
        raise ValueError(
            f"No existe un vendedor activo con el nombre exacto «{nombre}». "
            "Debe coincidir tal como está registrado en el sistema."
        )
    if len(rows) > 1:
        raise ValueError(
            f"Hay más de un vendedor activo con el nombre exacto «{nombre}»."
        )
    return rows[0]


def generar_plantilla_excel(db: Session) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"
    ws.append(PLANTILLA_HEADERS)
    ws.append(
        [
            "Empresa Ejemplo S.A.S.",
            "Contacto",
            "900123456-1",
            "Calle 1 # 2-3",
            "3001234567",
            "cliente@ejemplo.com",
            "Barranquilla",
            "",
        ]
    )
    ws_v = wb.create_sheet("Vendedores")
    ws_v.append(["nombre", "email"])
    for v in db.execute(
        select(Usuario)
        .where(Usuario.rol == RolUsuario.VENDEDOR.value, Usuario.activo.is_(True))
        .order_by(Usuario.nombre.asc())
    ).scalars():
        ws_v.append([v.nombre, v.email])
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _map_headers(header_row: tuple[Any, ...]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        key = _normalize_header(cell)
        if key in PLANTILLA_HEADERS and key not in mapping:
            mapping[key] = idx
    missing = [h for h in PLANTILLA_HEADERS if h not in mapping]
    if missing:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            f"Faltan columnas en la plantilla: {', '.join(missing)}. "
            f"Use la plantilla oficial con encabezados: {', '.join(PLANTILLA_HEADERS)}.",
        )
    return mapping


def _row_dict(values: tuple[Any, ...], col_map: dict[str, int]) -> dict[str, str]:
    out: dict[str, str] = {}
    for field in PLANTILLA_HEADERS:
        idx = col_map[field]
        out[field] = _cell_str(values[idx]) if idx < len(values) else ""
    return out


def importar_clientes_excel(db: Session, file: UploadFile) -> ClienteCargaMasivaResultado:
    filename = (file.filename or "").lower()
    if not filename.endswith((".xlsx", ".xlsm")):
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Formato no soportado. Suba un archivo Excel (.xlsx).",
        )
    raw = file.file.read()
    if not raw:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "El archivo está vacío.")
    if len(raw) > 5 * 1024 * 1024:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "El archivo supera el tamaño máximo permitido (5 MB).",
        )

    try:
        wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "No se pudo leer el archivo Excel.",
        ) from exc

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "El archivo no contiene filas de datos (solo encabezados o está vacío).",
        )

    col_map = _map_headers(rows[0])
    resultados: list[FilaCargaClienteResultado] = []
    nits_archivo: set[str] = set()
    creados = 0

    for row_num, row in enumerate(rows[1:], start=2):
        data = _row_dict(row, col_map)
        if not any(data.values()):
            continue

        nit = data["nit"]
        try:
            if not data["nombre"]:
                raise ValueError("El nombre es obligatorio.")
            if not nit:
                raise ValueError("El NIT es obligatorio.")
            if len(nit) < 3:
                raise ValueError("El NIT debe tener al menos 3 caracteres.")
            if nit in nits_archivo:
                raise ValueError(f"NIT duplicado en el archivo: {nit}")
            nits_archivo.add(nit)

            exists = db.execute(
                select(Cliente.id).where(Cliente.nit == nit)
            ).scalar_one_or_none()
            if exists:
                raise ValueError(f"Ya existe un cliente con NIT {nit}.")

            vendedor = _resolve_vendedor_exacto(db, data["vendedor"])
            correo = _parse_correo(data["correo"])

            cliente = Cliente(
                nombre=data["nombre"],
                apellidos=data["apellidos"] or None,
                nit=nit,
                direccion=data["direccion"] or None,
                telefono=data["telefono"] or None,
                correo=correo,
                ciudad=data["ciudad"] or None,
                vendedor_asignado_id=vendedor.id,
                activo=True,
            )
            db.add(cliente)
            db.flush()
            creados += 1
            resultados.append(
                FilaCargaClienteResultado(
                    fila=row_num,
                    nit=nit,
                    exito=True,
                    mensaje=f"Cliente creado y asignado a {vendedor.nombre}.",
                    cliente_id=cliente.id,
                )
            )
        except ValueError as exc:
            resultados.append(
                FilaCargaClienteResultado(
                    fila=row_num,
                    nit=nit or None,
                    exito=False,
                    mensaje=str(exc),
                )
            )

    db.commit()
    errores = sum(1 for r in resultados if not r.exito)
    return ClienteCargaMasivaResultado(
        total_filas=len(resultados),
        creados=creados,
        errores=errores,
        filas=resultados,
    )
